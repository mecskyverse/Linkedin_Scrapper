import scrapy
import pandas as pd
import openpyxl

class LinkedInPeopleProfileSpider(scrapy.Spider):
    name = "linkedin_people_profile"

    custom_settings = {
        'FEEDS': { 'data/%(name)s_%(time)s.jsonl': { 'format': 'jsonlines',}}
        }
    
    def fetch_links(self, path):
        try:
            # Read the CSV file
            df = pd.read_csv(path)

            # Check if the 'links' column exists
            if 'links' in df.columns:
                # Extract the links from the 'links' column
                links = df['links'].tolist()

                # Filter out empty or NaN values
                links = [link for link in links if pd.notna(link) and str(link).strip()]
                print("links=", links)
                
                if links:
                    return links
                else:
                    raise ValueError("No links to extract. Either crawl linkedin_website_scrape again or check your credits.")
            else:
                raise ValueError("CSV file does not contain a 'links' column.")

        except Exception as e:
            print(f"Error: {e}")
            # Handle the error or re-raise it as needed
    def write_linkedin_data(self, data, filename="linkedin_data.xlsx", sheet_name="Profile"):
        """
        Write LinkedIn profile data to an Excel file, creating a new sheet on the first call and appending data to existing rows on subsequent calls.

        Args:
            data: A dictionary containing the LinkedIn profile data.
            filename: The filename of the Excel file (default: "linkedin_data.xlsx").
            sheet_name: The name of the sheet within the Excel file (default: "Profile").
        """

        # Opening the workbook (creating it if it doesn't exist)
        try:
            # Trying to load the existing workbook
            wb = openpyxl.load_workbook(filename=filename, data_only=True)
        except FileNotFoundError:
            # If the file does not exist, creating a new workbook
            wb = openpyxl.Workbook()
            # Save the new workbook to the specified filename
            wb.save(filename)
        
        # Checking if the sheet exists
        if sheet_name not in wb.sheetnames:
            # Creating the sheet if it doesn't exist
            sheet = wb.create_sheet(sheet_name)

            # Writing header row if it's the first call
            sheet.append(["URL", "Name", "Description", "Location", "Followers", "Connections", "About"])
        else:
            sheet = wb["{}".format(sheet_name)]

        # Writing data to a new row
        sheet.append([
            data["url"],
            data["name"],
            data["description"],
            data["location"],
            data["followers"],
            data["connections"],
            data["about"],
        ])

        # Save the workbook
        wb.save(filename)



    def start_requests(self):
        url_list = self.fetch_links("output.csv")
        print("urlList=", url_list)
        for url in url_list:
            linkedin_people_url = url
            print('url= ',url) 
            yield scrapy.Request(url=linkedin_people_url, callback=self.parse_profile, meta={'linkedin_url': linkedin_people_url})

    def parse_profile(self, response):
        item = {}
        # item['profile'] = response.meta['profile']
        item['url'] = response.meta['linkedin_url']

        """
            SUMMARY SECTION
        """
        summary_box = response.css("section.top-card-layout")
        item['name'] = summary_box.css("h1::text").get().strip()
        item['description'] = summary_box.css("h2::text").get().strip()

        ## Location
        try:
            item['location'] = summary_box.css('div.top-card__subline-item::text').get()
        except:
            item['location'] = summary_box.css('span.top-card__subline-item::text').get().strip()
            if 'followers' in item['location'] or 'connections' in item['location']:
                item['location'] = ''

        item['followers'] = ''
        item['connections'] = ''

        for span_text in summary_box.css('span.top-card__subline-item::text').getall():
            if 'followers' in span_text:
                item['followers'] = span_text.replace(' followers', '').strip()
            if 'connections' in span_text:
                item['connections'] = span_text.replace(' connections', '').strip()


        """
            ABOUT SECTION
        """
        item['about'] = response.css('section.summary div.core-section-container__content p::text').get()


        # """
        #     EXPERIENCE SECTION
        # """
        # item['experience'] = []
        # experience_blocks = response.css('li.experience-item')
        # for block in experience_blocks:
        #     experience = {}
        #     ## organisation profile url
        #     try:
        #         experience['organisation_profile'] = block.css('h4 a::attr(href)').get().split('?')[0]
        #     except Exception as e:
        #         print('experience --> organisation_profile', e)
        #         experience['organisation_profile'] = ''
                
                
        #     ## location
        #     try:
        #         experience['location'] = block.css('p.experience-item__location::text').get().strip()
        #     except Exception as e:
        #         print('experience --> location', e)
        #         experience['location'] = ''
                
                
        #     ## description
        #     try:
        #         experience['description'] = block.css('p.show-more-less-text__text--more::text').get().strip()
        #     except Exception as e:
        #         print('experience --> description', e)
        #         try:
        #             experience['description'] = block.css('p.show-more-less-text__text--less::text').get().strip()
        #         except Exception as e:
        #             print('experience --> description', e)
        #             experience['description'] = ''
                    
        #     ## time range
        #     try:
        #         date_ranges = block.css('span.date-range time::text').getall()
        #         if len(date_ranges) == 2:
        #             experience['start_time'] = date_ranges[0]
        #             experience['end_time'] = date_ranges[1]
        #             experience['duration'] = block.css('span.date-range__duration::text').get()
        #         elif len(date_ranges) == 1:
        #             experience['start_time'] = date_ranges[0]
        #             experience['end_time'] = 'present'
        #             experience['duration'] = block.css('span.date-range__duration::text').get()
        #     except Exception as e:
        #         print('experience --> time ranges', e)
        #         experience['start_time'] = ''
        #         experience['end_time'] = ''
        #         experience['duration'] = ''
            
        #     item['experience'].append(experience)

        
        # """
        #     EDUCATION SECTION
        # """
        # item['education'] = []
        # education_blocks = response.css('li.education__list-item')
        # for block in education_blocks:
        #     education = {}

        #     ## organisation
        #     try:
        #         education['organisation'] = block.css('h3::text').get().strip()
        #     except Exception as e:
        #         print("education --> organisation", e)
        #         education['organisation'] = ''


        #     ## organisation profile url
        #     try:
        #         education['organisation_profile'] = block.css('a::attr(href)').get().split('?')[0]
        #     except Exception as e:
        #         print("education --> organisation_profile", e)
        #         education['organisation_profile'] = ''

        #     ## course details
        #     try:
        #         education['course_details'] = ''
        #         for text in block.css('h4 span::text').getall():
        #             education['course_details'] = education['course_details'] + text.strip() + ' '
        #         education['course_details'] = education['course_details'].strip()
        #     except Exception as e:
        #         print("education --> course_details", e)
        #         education['course_details'] = ''

        #     ## description
        #     try:
        #         education['description'] = block.css('div.education__item--details p::text').get().strip()
        #     except Exception as e:
        #         print("education --> description", e)
        #         education['description'] = ''

         
        #     ## time range
        #     try:
        #         date_ranges = block.css('span.date-range time::text').getall()
        #         if len(date_ranges) == 2:
        #             education['start_time'] = date_ranges[0]
        #             education['end_time'] = date_ranges[1]
        #         elif len(date_ranges) == 1:
        #             education['start_time'] = date_ranges[0]
        #             education['end_time'] = 'present'
        #     except Exception as e:
        #         print("education --> time_ranges", e)
        #         education['start_time'] = ''
        #         education['end_time'] = ''

        #     item['education'].append(education)
       
        """
            GEN AI SECTION
        """
        # not much of data so currently no need to include GEN AI for the filtration as we can already get filtered data from scrapping


        print("final Item", item)
        self.write_linkedin_data(item)
        yield item
        
    
